from openpyxl import load_workbook
import formParser


def listToExcel(dataList, workBookName):

    wb = load_workbook(workBookName)
    ws = wb.active

    dynamic_hash = createHashTypeDinamiclly(ws)

    ws.title = "form name"
    row_to_insert = ws.max_row+1
    for header, data_input in dataList:
        this_column = dynamic_hash.get(header, -1)
        if(this_column != -1):
            ws.cell(row=row_to_insert, column=this_column, value=data_input)

    wb.save(workBookName)
    wb.close()


def createHashTypeDinamiclly(ws):
    dynamic_hash_type = {}
    TITLE_ROW = 1
    for title in ws.iter_cols(min_row=TITLE_ROW, max_col=ws.max_column, max_row=1, ):
        key = title[0].value
        val = title[0].column
        dynamic_hash_type[key] = val

    return dynamic_hash_type
